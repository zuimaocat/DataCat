import requests,openpyxl,re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from tkinter import PhotoImage
import tkinter as tk
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# https://www.chinawutong.com/201/s11883/   网站


headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'}
wb = Workbook()

def start_spider(base_url):

    heavy_prices = []   
    light_prices = []
    company_names = []
    company_vips = []
    values = [company_names, heavy_prices, light_prices, company_vips]
    
    base_res = requests.get(base_url, headers=headers)
    base_soup = BeautifulSoup(base_res.text, 'html.parser')

    #获取页数信息
    page  = int(base_soup.find_all('span',class_='p_a',)[-1].text)
    
    
    def sub_spider(url):
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            price_elements = soup.select('div.priceZxInfo')
            company_elements = soup.select('div.clearFix')
            vips = soup.find_all('li',class_='tuiJianListYear')

                
            prices = []
            for i in range(len(price_elements)):
                # 提取文本内容
                
                price_text = price_elements[i].get_text(strip=True)
                
                # 查找公司名称
                company_name = company_elements[i].find("h2")
                
                #查找公司会员
                if i < len(vips):
                    company_vip = int(vips[i].find('span',class_='fontOrange').text)
                else:
                    company_vip = 0
                if '元/' in price_text and company_name:
                    prices.append(price_text)
                    company_names.append(company_name.get_text(strip=True))
                    company_vips.append(company_vip)

            
        
            for i in range(len(prices)):
                # 使用正则表达式匹配价格部分，然后提取数字
                heavy_match = re.search(r'重货：(\d+(\.\d+)?)', prices[i])
                light_match = re.search(r'轻货：(\d+(\.\d+)?)', prices[i])

                if heavy_match:
                    heavy_price = float(heavy_match.group(1))
                else:
                    heavy_price = None

                if light_match:
                    light_price = float(light_match.group(1))
                else:
                    light_price = None
                if heavy_price < 10:  #单位转换
                    heavy_price *= 1000
                heavy_prices.append(heavy_price)
                light_prices.append(light_price)
            print("已完成数据爬取,准备写入\n")
    for p in range(1,page):
        url = base_url + '?p=' + str(p)
        sub_spider(url)
    #写入excel
    def write_excel(place):
        # global company_names, heavy_prices, light_prices, company_vips
        
        # 创建新工作表
        new_sheet = wb.create_sheet(title=place)
        sheet = wb.get_sheet_by_name(place)
        line1 = ['公司', '重货价.元/吨', '轻货价.元/方', '资历.年']
        num = 1
        for i in line1:  #写入标题
            sheet.cell(1, num, i)
            num += 1
        
        # 将数据写入excel
        row = 1
        for value in values:
            line = 2
            for i in value:
                sheet.cell(line, row, i)
                line += 1
            row += 1
        
        
        # 将生成器对象转换为列表
        rows = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row))
        
        # 反向遍历列表
        for row in reversed(rows):
            # 检查第一列是否为空
            if row[0].value is None:
                # 删除该行及其之后的所有内容（包括第四列的内容）
                sheet.delete_rows(row[0].row, amount=sheet.max_row - row[0].row + 1)
        
        print('已完成全部爬取,并生成了excel表格')
        print(company_vips,heavy_prices,light_prices,company_names,company_vips)
    #获取地点信息 
    place_text = base_soup.find('div',class_='leftZhuanXian_l').find('a').text
    place = base_soup.find('div',class_='leftZhuanXian_l').find('a').text[:place_text.find('物流')]
    write_excel(place)
    
    wb.save('公司货流数据.xlsx')

def start_gui():

    #链接爬虫
    def multi_spider_url():
        urls_text = text.get("1.0","end-1c")
        urls = urls_text.split('\n')

        for url in urls:
            print(url)
            start_spider(url)

    #地点名爬虫
    def name_spider():
        driver = webdriver.Chrome() 
        driver.get("https://www.chinawutong.com/201/")
        sheng_from = text3.get("1.0","end-1c")+'省'
        city_from = text4.get("1.0","end-1c")+"市"
        sheng_to = text_end_sheng.get("1.0","end-1c")+'省'
        city_to = text_end_shi.get("1.0","end-1c")+"市"
        driver.maximize_window()
        driver.find_element(By.CSS_SELECTOR,"#tbFrom").click()

        religion = driver.find_element(By.CSS_SELECTOR, 'span.wtm_a_txt[txt={}]'.format(sheng_from)).click()
        time.sleep(1)

        religion = driver.find_element(By.CSS_SELECTOR, 'span.wtm_a_txt[txt={}]'.format(city_from)).click()

        driver.find_element(By.CSS_SELECTOR,"#tbTo").click()
        religion = driver.find_element(By.CSS_SELECTOR, 'span.wtm_a_txt[txt={}]'.format(sheng_to)).click()
        time.sleep(1)


        religion = driver.find_element(By.CSS_SELECTOR, 'span.wtm_a_txt[txt={}]'.format(city_to)).click()
        driver.find_element(By.CSS_SELECTOR, 'span.orangeBtn').click()
        time.sleep(1)
        url = driver.current_url
        start_spider(url)


    root = tk.Tk()
    root.title("全域物流方案规划系统|货流数据挖掘工具V2.1.1")
    root.iconbitmap("logo.ico")
    root.geometry('440x330')
    logo = PhotoImage(file="logo.png")
    logo_label = tk.Label(root, image=logo)
    logo_label.place(x=1,y=1)
    text = tk.Text(root, width= 30, height= 7)

    label2 = tk.Label(root,text='请输入需要获取数据的链接（可输入多个链接）')
    label_split = tk.Label(root,text='————————————————————————————')

    button2 = tk.Button(root,text = '开始',command=multi_spider_url)
    label3 = tk.Label(root,text='请输入起点和终点')
    text3 = tk.Text(root, width= 5, height= 1)
    label5 = tk.Label(root,text='省')
    text4 = tk.Text(root, width= 5, height= 1)
    label6 = tk.Label(root,text='市')
    label_TO = tk.Label(root,text='到')

    text_end_sheng = tk.Text(root, width= 5, height= 1)
    label_end_sheng = tk.Label(root,text='省')
    text_end_shi= tk.Text(root, width= 5, height= 1)
    label_end_shi = tk.Label(root,text='市')

    text_end_sheng.place(x=210, y=250)
    label_end_sheng.place(x=250, y=250)
    text_end_shi.place(x=280, y=250)
    label_end_shi.place(x=320, y=250)
    button_text = tk.Button(root,text = '开始',command=name_spider)

    button_text.place(x=350,y=250)
    label2.pack()
    text.pack()
    button2.pack()

    label3.place(x=10,y=220)
    text3.place(x=15, y=250)
    text4.place(x=80, y=250)
    label5.place(x=50, y=250)
    label6.place(x=120, y=250)
    label_TO.place(x=160,y=250)
    
    
    root.mainloop()
    
start_gui()




